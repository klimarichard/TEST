import os
import click
import requests
from io import BytesIO
from datetime import datetime, date
from functools import wraps
from flask import (Flask, render_template, redirect, url_for, request,
                   flash, jsonify, abort, send_file)
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from flask_sqlalchemy import SQLAlchemy
from flask_migrate import Migrate
from flask_login import (LoginManager, UserMixin, login_user, logout_user,
                         login_required, current_user)
from flask_wtf.csrf import CSRFProtect
from flask_mail import Mail, Message
from werkzeug.security import generate_password_hash, check_password_hash

# ─────────────────────────────────────────────────────────────────────────────
# App setup
# ─────────────────────────────────────────────────────────────────────────────

app = Flask(__name__)
app.config['SECRET_KEY'] = os.environ.get('SECRET_KEY', 'change-this-secret-key-in-production')

# Railway provides postgres:// but SQLAlchemy requires postgresql://
_db_url = os.environ.get('DATABASE_URL', 'sqlite:///carexpenses.db')
if _db_url.startswith('postgres://'):
    _db_url = _db_url.replace('postgres://', 'postgresql://', 1)
app.config['SQLALCHEMY_DATABASE_URI'] = _db_url
app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = False

db = SQLAlchemy(app)
migrate = Migrate(app, db)
csrf = CSRFProtect(app)
login_manager = LoginManager(app)
login_manager.login_view = 'login'
login_manager.login_message_category = 'warning'

app.config['MAIL_SERVER']         = os.environ.get('MAIL_SERVER', 'smtp.example.com')
app.config['MAIL_PORT']           = int(os.environ.get('MAIL_PORT', 587))
app.config['MAIL_USE_TLS']        = os.environ.get('MAIL_USE_TLS', 'true').lower() == 'true'
app.config['MAIL_USERNAME']       = os.environ.get('MAIL_USERNAME')
app.config['MAIL_PASSWORD']       = os.environ.get('MAIL_PASSWORD')
app.config['MAIL_DEFAULT_SENDER'] = os.environ.get('MAIL_DEFAULT_SENDER', 'noreply@carexpenses.app')
mail = Mail(app)


# ─────────────────────────────────────────────────────────────────────────────
# Models
# ─────────────────────────────────────────────────────────────────────────────

class User(UserMixin, db.Model):
    __tablename__ = 'users'
    id = db.Column(db.Integer, primary_key=True)
    username = db.Column(db.String(80), unique=True, nullable=False)
    email = db.Column(db.String(120), unique=True, nullable=False)
    password_hash = db.Column(db.String(256), nullable=False)
    is_admin = db.Column(db.Boolean, default=False)
    theme = db.Column(db.String(10), default='light', nullable=False)
    pending_username = db.Column(db.String(80), nullable=True)
    pending_username_at = db.Column(db.DateTime, nullable=True)
    created_at = db.Column(db.DateTime, default=datetime.utcnow)

    owned_cars = db.relationship('Car', backref='owner', lazy=True)
    shared_cars = db.relationship('CarShare', backref='user', lazy=True)
    expenses = db.relationship('Expense', backref='user', lazy=True)

    def set_password(self, password):
        self.password_hash = generate_password_hash(password)

    def check_password(self, password):
        return check_password_hash(self.password_hash, password)

    def get_accessible_cars(self):
        owned = Car.query.filter_by(owner_id=self.id).all()
        shared_ids = [s.car_id for s in self.shared_cars]
        shared = Car.query.filter(Car.id.in_(shared_ids)).all() if shared_ids else []
        seen, result = set(), []
        for c in owned + shared:
            if c.id not in seen:
                seen.add(c.id)
                result.append(c)
        return result


class Car(db.Model):
    __tablename__ = 'cars'
    id = db.Column(db.Integer, primary_key=True)
    name = db.Column(db.String(100), nullable=False)
    make = db.Column(db.String(80))
    model_name = db.Column(db.String(80))
    year = db.Column(db.Integer)
    license_plate = db.Column(db.String(20))
    owner_id = db.Column(db.Integer, db.ForeignKey('users.id'), nullable=False)
    created_at = db.Column(db.DateTime, default=datetime.utcnow)

    shares = db.relationship('CarShare', backref='car', lazy=True, cascade='all, delete-orphan')
    expenses = db.relationship('Expense', backref='car', lazy=True, cascade='all, delete-orphan')

    def is_accessible_by(self, user):
        return (self.owner_id == user.id or
                CarShare.query.filter_by(car_id=self.id, user_id=user.id).first() is not None)

    @property
    def display_name(self):
        parts = [self.name]
        if self.make and self.model_name:
            parts.append(f'({self.make} {self.model_name})')
        elif self.make:
            parts.append(f'({self.make})')
        return ' '.join(parts)


class CarShare(db.Model):
    __tablename__ = 'car_shares'
    id = db.Column(db.Integer, primary_key=True)
    car_id = db.Column(db.Integer, db.ForeignKey('cars.id'), nullable=False)
    user_id = db.Column(db.Integer, db.ForeignKey('users.id'), nullable=False)
    __table_args__ = (db.UniqueConstraint('car_id', 'user_id'),)


EXPENSE_TYPES = ['fuel', 'repair', 'toll', 'insurance', 'gadget']
TYPE_LABELS = {
    'fuel': 'Fuel', 'repair': 'Repair', 'toll': 'Toll',
    'insurance': 'Insurance', 'gadget': 'Gadget',
}
TYPE_ICONS = {
    'fuel': 'bi-fuel-pump', 'repair': 'bi-wrench-adjustable',
    'toll': 'bi-signpost-2', 'insurance': 'bi-shield-check',
    'gadget': 'bi-phone',
}
TYPE_COLORS = {
    'fuel': '#0dcaf0', 'repair': '#dc3545', 'toll': '#ffc107',
    'insurance': '#198754', 'gadget': '#6f42c1',
}


class Expense(db.Model):
    __tablename__ = 'expenses'
    id = db.Column(db.Integer, primary_key=True)
    car_id = db.Column(db.Integer, db.ForeignKey('cars.id'), nullable=False)
    user_id = db.Column(db.Integer, db.ForeignKey('users.id'), nullable=False)
    date = db.Column(db.Date, nullable=False)
    amount = db.Column(db.Float, nullable=False)
    currency = db.Column(db.String(3), default='CZK')
    amount_czk = db.Column(db.Float, nullable=False)
    expense_type = db.Column(db.String(20), nullable=False)
    notes = db.Column(db.Text)
    created_at = db.Column(db.DateTime, default=datetime.utcnow)

    fuel_detail = db.relationship('FuelDetail', backref='expense', uselist=False, cascade='all, delete-orphan')
    repair_detail = db.relationship('RepairDetail', backref='expense', uselist=False, cascade='all, delete-orphan')
    toll_detail = db.relationship('TollDetail', backref='expense', uselist=False, cascade='all, delete-orphan')
    insurance_detail = db.relationship('InsuranceDetail', backref='expense', uselist=False, cascade='all, delete-orphan')
    gadget_detail = db.relationship('GadgetDetail', backref='expense', uselist=False, cascade='all, delete-orphan')

    def can_edit(self, user):
        return user.is_admin or self.user_id == user.id or self.car.owner_id == user.id

    @property
    def type_label(self):
        return TYPE_LABELS.get(self.expense_type, self.expense_type.title())

    @property
    def type_icon(self):
        return TYPE_ICONS.get(self.expense_type, 'bi-cash')

    @property
    def type_color(self):
        return TYPE_COLORS.get(self.expense_type, '#6c757d')

    @property
    def detail(self):
        return (self.fuel_detail or self.repair_detail or
                self.toll_detail or self.insurance_detail or self.gadget_detail)


class FuelDetail(db.Model):
    __tablename__ = 'fuel_details'
    id = db.Column(db.Integer, primary_key=True)
    expense_id = db.Column(db.Integer, db.ForeignKey('expenses.id'), nullable=False)
    liters = db.Column(db.Float)
    price_per_liter = db.Column(db.Float)
    odometer = db.Column(db.Integer)
    station_location = db.Column(db.String(200))
    transaction_time = db.Column(db.Time)

    @property
    def summary(self):
        parts = []
        if self.liters:
            parts.append(f'{self.liters:.1f} L')
        if self.station_location:
            parts.append(self.station_location)
        if self.odometer:
            parts.append(f'@ {self.odometer:,} km')
        return ' · '.join(parts) if parts else '—'


class RepairDetail(db.Model):
    __tablename__ = 'repair_details'
    id = db.Column(db.Integer, primary_key=True)
    expense_id = db.Column(db.Integer, db.ForeignKey('expenses.id'), nullable=False)
    description = db.Column(db.Text)

    @property
    def summary(self):
        return (self.description or '—')[:80]


class TollDetail(db.Model):
    __tablename__ = 'toll_details'
    id = db.Column(db.Integer, primary_key=True)
    expense_id = db.Column(db.Integer, db.ForeignKey('expenses.id'), nullable=False)
    route = db.Column(db.String(200))
    toll_type = db.Column(db.String(50))
    expiration_date = db.Column(db.Date, nullable=True)
    notify_before_expiry = db.Column(db.Boolean, default=False)
    notify_days_before = db.Column(db.Integer, default=14)
    notification_sent = db.Column(db.Boolean, default=False)

    @property
    def summary(self):
        parts = []
        if self.toll_type:
            parts.append(self.toll_type)
        if self.route:
            parts.append(self.route)
        if self.expiration_date:
            parts.append(f'exp. {self.expiration_date.strftime("%d.%m.%Y")}')
        return ' · '.join(parts) if parts else '—'

    @property
    def is_expiring_soon(self):
        if not self.expiration_date:
            return False
        from datetime import timedelta
        return date.today() < self.expiration_date <= date.today() + timedelta(days=30)

    @property
    def is_expired(self):
        return bool(self.expiration_date and self.expiration_date < date.today())


class InsuranceDetail(db.Model):
    __tablename__ = 'insurance_details'
    id = db.Column(db.Integer, primary_key=True)
    expense_id = db.Column(db.Integer, db.ForeignKey('expenses.id'), nullable=False)
    insurance_type = db.Column(db.String(100))
    provider = db.Column(db.String(100))
    expiration_date = db.Column(db.Date)
    notify_before_expiry = db.Column(db.Boolean, default=False)
    notify_days_before = db.Column(db.Integer, default=14)
    notification_sent = db.Column(db.Boolean, default=False)

    @property
    def summary(self):
        parts = []
        if self.insurance_type:
            parts.append(self.insurance_type)
        if self.provider:
            parts.append(self.provider)
        if self.expiration_date:
            parts.append(f'exp. {self.expiration_date.strftime("%d.%m.%Y")}')
        return ' · '.join(parts) if parts else '—'

    @property
    def is_expiring_soon(self):
        if not self.expiration_date:
            return False
        from datetime import timedelta
        return date.today() < self.expiration_date <= date.today() + timedelta(days=30)

    @property
    def is_expired(self):
        return bool(self.expiration_date and self.expiration_date < date.today())


class GadgetDetail(db.Model):
    __tablename__ = 'gadget_details'
    id = db.Column(db.Integer, primary_key=True)
    expense_id = db.Column(db.Integer, db.ForeignKey('expenses.id'), nullable=False)
    gadget_type = db.Column(db.String(100))

    @property
    def summary(self):
        return self.gadget_type or '—'


@login_manager.user_loader
def load_user(user_id):
    return User.query.get(int(user_id))


# ─────────────────────────────────────────────────────────────────────────────
# Helpers
# ─────────────────────────────────────────────────────────────────────────────

def admin_required(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        if not current_user.is_authenticated:
            return redirect(url_for('login', next=request.url))
        if not current_user.is_admin:
            abort(403)
        return f(*args, **kwargs)
    return decorated


_rate_cache: dict = {}


def get_exchange_rate(from_currency: str, to_currency: str = 'CZK') -> float | None:
    if from_currency == to_currency:
        return 1.0
    cache_key = f'{from_currency}_{to_currency}'
    cached = _rate_cache.get(cache_key)
    if cached:
        ts, rate = cached
        if (datetime.utcnow() - ts).seconds < 3600:
            return rate
    try:
        r = requests.get(
            'https://api.frankfurter.app/latest',
            params={'from': from_currency, 'to': to_currency},
            timeout=5,
        )
        r.raise_for_status()
        rate = r.json()['rates'][to_currency]
        _rate_cache[cache_key] = (datetime.utcnow(), rate)
        return rate
    except Exception:
        return None


SUPPORTED_CURRENCIES = ['CZK', 'EUR', 'USD', 'GBP', 'CHF', 'PLN', 'HUF', 'NOK', 'SEK', 'DKK']
TOLL_TYPES = ['Highway vignette', 'Toll gate', 'Bridge toll', 'Tunnel toll', 'Parking', 'Other']
INSURANCE_TYPES = ['Liability (Povinné ručení)', 'Comprehensive (Havarijní)', 'GAP Insurance', 'Other']


@app.template_filter('czk')
def format_czk(value):
    if value is None:
        return '—'
    formatted = f'{value:,.0f}'.replace(',', '\u202f')  # narrow no-break space (CZ thousands sep)
    return f'{formatted}\u00a0Kč'


def _export_xlsx(expenses, include_user=False, filename='expenses'):
    """Build and return an xlsx response for a list of expenses."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Expenses'

    header_font = Font(bold=True)
    header_fill = PatternFill('solid', fgColor='D9E1F2')

    headers = ['Date', 'Time', 'Car']
    if include_user:
        headers.append('User')
    headers += ['Type', 'Details', 'Amount (CZK)', 'Orig. Amount', 'Currency', 'Notes']
    ws.append(headers)
    for col, cell in enumerate(ws[1], 1):
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')

    for e in expenses:
        time_str = ''
        if e.expense_type == 'fuel' and e.fuel_detail and e.fuel_detail.transaction_time:
            time_str = e.fuel_detail.transaction_time.strftime('%H:%M')
        detail_str = e.detail.summary if e.detail else ''
        row = [e.date.strftime('%d.%m.%Y'), time_str, e.car.name]
        if include_user:
            row.append(e.user.username)
        row += [
            TYPE_LABELS.get(e.expense_type, e.expense_type),
            detail_str,
            round(e.amount_czk, 2),
            e.amount if e.currency != 'CZK' else '',
            e.currency if e.currency != 'CZK' else '',
            e.notes or '',
        ]
        ws.append(row)

    # Auto-fit column widths (approximate)
    for col in ws.columns:
        max_len = max((len(str(cell.value or '')) for cell in col), default=8)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 50)

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return send_file(buf, as_attachment=True,
                     download_name=f'{filename}.xlsx',
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')


@app.context_processor
def inject_globals():
    pending_count = 0
    if current_user.is_authenticated and current_user.is_admin:
        pending_count = User.query.filter(User.pending_username.isnot(None)).count()
    return {
        'now': datetime.utcnow(),
        'TYPE_COLORS': TYPE_COLORS,
        'TYPE_LABELS': TYPE_LABELS,
        'TYPE_ICONS': TYPE_ICONS,
        'EXPENSE_TYPES': EXPENSE_TYPES,
        'pending_username_count': pending_count,
    }


# ─────────────────────────────────────────────────────────────────────────────
# Error handlers
# ─────────────────────────────────────────────────────────────────────────────

@app.errorhandler(403)
def forbidden(e):
    return render_template('errors/403.html'), 403


@app.errorhandler(404)
def not_found(e):
    return render_template('errors/404.html'), 404


# ─────────────────────────────────────────────────────────────────────────────
# Auth
# ─────────────────────────────────────────────────────────────────────────────

@app.route('/')
def index():
    return redirect(url_for('dashboard') if current_user.is_authenticated else url_for('login'))


@app.route('/setup', methods=['GET', 'POST'])
def setup():
    if User.query.first():
        return redirect(url_for('login'))
    if request.method == 'POST':
        username = request.form.get('username', '').strip()
        email = request.form.get('email', '').strip()
        password = request.form.get('password', '')
        confirm = request.form.get('confirm_password', '')
        if not username or not email or not password:
            flash('All fields are required.', 'danger')
        elif password != confirm:
            flash('Passwords do not match.', 'danger')
        else:
            user = User(username=username, email=email, is_admin=True)
            user.set_password(password)
            db.session.add(user)
            db.session.commit()
            flash('Admin account created. Please sign in.', 'success')
            return redirect(url_for('login'))
    return render_template('setup.html')


@app.route('/login', methods=['GET', 'POST'])
def login():
    if current_user.is_authenticated:
        return redirect(url_for('dashboard'))
    if not User.query.first():
        return redirect(url_for('setup'))
    if request.method == 'POST':
        username = request.form.get('username', '').strip()
        password = request.form.get('password', '')
        user = User.query.filter_by(username=username).first()
        if user and user.check_password(password):
            login_user(user, remember=bool(request.form.get('remember')))
            return redirect(request.args.get('next') or url_for('dashboard'))
        flash('Invalid username or password.', 'danger')
    return render_template('login.html')


@app.route('/logout')
@login_required
def logout():
    logout_user()
    flash('You have been logged out.', 'info')
    return redirect(url_for('login'))


@app.route('/account', methods=['GET', 'POST'])
@login_required
def account():
    if request.method == 'POST':
        action = request.form.get('action')

        if action == 'update_email':
            email = request.form.get('email', '').strip()
            if not email:
                flash('Email is required.', 'danger')
            elif User.query.filter(User.email == email, User.id != current_user.id).first():
                flash('That email is already in use.', 'danger')
            else:
                current_user.email = email
                db.session.commit()
                flash('Email updated.', 'success')

        elif action == 'update_password':
            current_pw = request.form.get('current_password', '')
            new_pw = request.form.get('new_password', '')
            confirm_pw = request.form.get('confirm_password', '')
            if not current_user.check_password(current_pw):
                flash('Current password is incorrect.', 'danger')
            elif len(new_pw) < 6:
                flash('New password must be at least 6 characters.', 'danger')
            elif new_pw != confirm_pw:
                flash('Passwords do not match.', 'danger')
            else:
                current_user.set_password(new_pw)
                db.session.commit()
                flash('Password updated.', 'success')

        elif action == 'request_username':
            new_username = request.form.get('new_username', '').strip()
            if not new_username:
                flash('Username is required.', 'danger')
            elif new_username == current_user.username:
                flash('That is already your username.', 'warning')
            elif User.query.filter(
                (User.username == new_username) | (User.pending_username == new_username)
            ).filter(User.id != current_user.id).first():
                flash('That username is already taken or pending approval.', 'danger')
            else:
                current_user.pending_username = new_username
                current_user.pending_username_at = datetime.utcnow()
                db.session.commit()
                flash('Username change request submitted. Awaiting admin approval.', 'info')

        elif action == 'cancel_username_request':
            current_user.pending_username = None
            current_user.pending_username_at = None
            db.session.commit()
            flash('Username change request cancelled.', 'info')

        return redirect(url_for('account'))
    return render_template('account.html')


@app.route('/account/theme', methods=['POST'])
@login_required
def account_theme():
    theme = request.form.get('theme', 'light')
    if theme in ('light', 'dark'):
        current_user.theme = theme
        db.session.commit()
    return ('', 204)


# ─────────────────────────────────────────────────────────────────────────────
# Dashboard
# ─────────────────────────────────────────────────────────────────────────────

@app.route('/dashboard')
@login_required
def dashboard():
    if current_user.is_admin:
        return redirect(url_for('admin_overview'))
    from collections import defaultdict
    cars = current_user.get_accessible_cars()
    car_ids = [c.id for c in cars]

    all_expenses = (Expense.query
                    .filter(Expense.car_id.in_(car_ids))
                    .order_by(Expense.date)
                    .all()) if car_ids else []

    car_totals = {c.id: 0.0 for c in cars}
    monthly_data: dict = defaultdict(float)
    type_totals: dict = defaultdict(float)

    for e in all_expenses:
        car_totals[e.car_id] += e.amount_czk
        monthly_data[e.date.strftime('%Y-%m')] += e.amount_czk
        type_totals[e.expense_type] += e.amount_czk

    sorted_months = sorted(monthly_data.keys())[-12:]
    this_month = datetime.utcnow().strftime('%Y-%m')
    this_month_total = monthly_data.get(this_month, 0.0)
    recent_expenses = sorted(all_expenses, key=lambda e: (e.date, e.created_at), reverse=True)[:10]

    return render_template('dashboard.html',
        cars=cars,
        car_totals=car_totals,
        recent_expenses=recent_expenses,
        total_spent=sum(type_totals.values()),
        total_expense_count=len(all_expenses),
        this_month_total=this_month_total,
        chart_monthly={
            'labels': sorted_months,
            'data': [monthly_data[m] for m in sorted_months],
        },
        chart_types={
            'labels': [TYPE_LABELS.get(t, t) for t in type_totals],
            'data': list(type_totals.values()),
            'colors': [TYPE_COLORS.get(t, '#6c757d') for t in type_totals],
        },
    )


# ─────────────────────────────────────────────────────────────────────────────
# User expense list
# ─────────────────────────────────────────────────────────────────────────────

@app.route('/expenses')
@login_required
def expense_list():
    if current_user.is_admin:
        return redirect(url_for('admin_overview'))
    from collections import defaultdict
    cars = current_user.get_accessible_cars()
    sel_car_ids = request.args.getlist('car_id', type=int)
    sel_types   = [t for t in request.args.getlist('type') if t in EXPENSE_TYPES]

    car_ids = [c.id for c in cars]
    q = Expense.query.filter(Expense.car_id.in_(car_ids)) if car_ids else Expense.query.filter(False)
    if sel_car_ids:
        q = q.filter(Expense.car_id.in_(sel_car_ids))
    if sel_types:
        q = q.filter(Expense.expense_type.in_(sel_types))
    expenses = q.order_by(Expense.date.desc(), Expense.created_at.desc()).all()

    type_totals: dict = defaultdict(float)
    for e in expenses:
        type_totals[e.expense_type] += e.amount_czk

    return render_template('expenses/list.html',
        cars=cars,
        expenses=expenses,
        sel_car_ids=sel_car_ids,
        sel_types=sel_types,
        total_czk=sum(e.amount_czk for e in expenses),
        type_totals=dict(type_totals),
    )


@app.route('/expenses/export')
@login_required
def expense_export():
    if current_user.is_admin:
        return redirect(url_for('admin_overview_export'))
    cars = current_user.get_accessible_cars()
    sel_car_ids = request.args.getlist('car_id', type=int)
    sel_types   = [t for t in request.args.getlist('type') if t in EXPENSE_TYPES]
    car_ids = [c.id for c in cars]
    q = Expense.query.filter(Expense.car_id.in_(car_ids)) if car_ids else Expense.query.filter(False)
    if sel_car_ids:
        q = q.filter(Expense.car_id.in_(sel_car_ids))
    if sel_types:
        q = q.filter(Expense.expense_type.in_(sel_types))
    expenses = q.order_by(Expense.date.desc(), Expense.created_at.desc()).all()
    return _export_xlsx(expenses, include_user=False, filename='my_expenses')


# ─────────────────────────────────────────────────────────────────────────────
# Cars
# ─────────────────────────────────────────────────────────────────────────────

@app.route('/cars')
@login_required
def cars_list():
    if current_user.is_admin:
        return redirect(url_for('admin_overview'))
    cars = current_user.get_accessible_cars()
    return render_template('cars/list.html', cars=cars)


@app.route('/cars/new', methods=['GET', 'POST'])
@login_required
def car_new():
    if current_user.is_admin:
        flash('Admin accounts cannot own cars.', 'warning')
        return redirect(url_for('admin_overview'))
    if request.method == 'POST':
        name = request.form.get('name', '').strip()
        if not name:
            flash('Car name is required.', 'danger')
            return render_template('cars/form.html', car=None)
        car = Car(
            name=name,
            make=request.form.get('make', '').strip() or None,
            model_name=request.form.get('model_name', '').strip() or None,
            year=request.form.get('year') or None,
            license_plate=request.form.get('license_plate', '').strip() or None,
            owner_id=current_user.id,
        )
        db.session.add(car)
        db.session.commit()
        flash(f'Car "{car.name}" added.', 'success')
        return redirect(url_for('car_detail', car_id=car.id))
    return render_template('cars/form.html', car=None)


@app.route('/cars/<int:car_id>')
@login_required
def car_detail(car_id):
    from collections import defaultdict
    car = Car.query.get_or_404(car_id)
    if not current_user.is_admin and not car.is_accessible_by(current_user):
        abort(403)
    is_admin_view = current_user.is_admin and car.owner_id != current_user.id

    expenses = (Expense.query
                .filter_by(car_id=car_id)
                .order_by(Expense.date.desc(), Expense.created_at.desc())
                .all())

    monthly_by_type: dict = defaultdict(lambda: defaultdict(float))
    type_totals: dict = defaultdict(float)
    for e in expenses:
        key = e.date.strftime('%Y-%m')
        monthly_by_type[key][e.expense_type] += e.amount_czk
        type_totals[e.expense_type] += e.amount_czk

    all_months = sorted(monthly_by_type.keys())[-12:]

    # Fuel efficiency date filter
    eff_from_str = request.args.get('eff_from', '').strip()
    eff_to_str   = request.args.get('eff_to', '').strip()
    try:
        eff_from = datetime.strptime(eff_from_str, '%Y-%m-%d').date() if eff_from_str else None
    except ValueError:
        eff_from = None
    try:
        eff_to = datetime.strptime(eff_to_str, '%Y-%m-%d').date() if eff_to_str else None
    except ValueError:
        eff_to = None

    # Fuel efficiency L/100 km
    fuel_exps = sorted(
        [e for e in expenses
         if e.expense_type == 'fuel' and e.fuel_detail
         and e.fuel_detail.odometer and e.fuel_detail.liters
         and (eff_from is None or e.date >= eff_from)
         and (eff_to   is None or e.date <= eff_to)],
        key=lambda e: e.fuel_detail.odometer,
    )
    eff_labels, eff_data = [], []
    for i in range(1, len(fuel_exps)):
        dist = fuel_exps[i].fuel_detail.odometer - fuel_exps[i - 1].fuel_detail.odometer
        if dist > 0:
            eff = fuel_exps[i].fuel_detail.liters / dist * 100
            eff_labels.append(fuel_exps[i].date.strftime('%d.%m.%Y'))
            eff_data.append(round(eff, 2))

    # Overall efficiency: total liters (fills 2..n) / total km
    overall_eff = None
    if len(fuel_exps) >= 2:
        total_liters = sum(e.fuel_detail.liters for e in fuel_exps[1:])
        total_km = fuel_exps[-1].fuel_detail.odometer - fuel_exps[0].fuel_detail.odometer
        if total_km > 0:
            overall_eff = round(total_liters / total_km * 100, 2)

    # Insurance warnings
    insurance_warnings = [
        e for e in expenses
        if e.expense_type == 'insurance' and e.insurance_detail
        and (e.insurance_detail.is_expired or e.insurance_detail.is_expiring_soon)
    ]
    toll_warnings = [
        e for e in expenses
        if e.expense_type == 'toll' and e.toll_detail and e.toll_detail.expiration_date
        and (e.toll_detail.is_expired or e.toll_detail.is_expiring_soon)
    ]

    return render_template('cars/detail.html',
        car=car,
        expenses=expenses,
        is_owner=(current_user.is_admin or car.owner_id == current_user.id),
        is_admin_view=is_admin_view,
        total_spent=sum(e.amount_czk for e in expenses),
        insurance_warnings=insurance_warnings,
        toll_warnings=toll_warnings,
        eff_from=eff_from_str,
        eff_to=eff_to_str,
        overall_eff=overall_eff,
        chart_stacked={
            'labels': all_months,
            'types': EXPENSE_TYPES,
            'data': {t: [monthly_by_type[m].get(t, 0) for m in all_months] for t in EXPENSE_TYPES},
        },
        chart_types={
            'labels': [TYPE_LABELS.get(t, t) for t in type_totals],
            'data': list(type_totals.values()),
            'colors': [TYPE_COLORS.get(t, '#6c757d') for t in type_totals],
        },
        chart_fuel_eff={'labels': eff_labels, 'data': eff_data},
    )


@app.route('/cars/<int:car_id>/edit', methods=['GET', 'POST'])
@login_required
def car_edit(car_id):
    car = Car.query.get_or_404(car_id)
    if not current_user.is_admin and car.owner_id != current_user.id:
        abort(403)
    if current_user.is_admin and car.owner_id != current_user.id:
        flash(f'Admin override: editing car owned by {car.owner.username}.', 'warning')
    if request.method == 'POST':
        name = request.form.get('name', '').strip()
        if not name:
            flash('Car name is required.', 'danger')
            return render_template('cars/form.html', car=car)
        car.name = name
        car.make = request.form.get('make', '').strip() or None
        car.model_name = request.form.get('model_name', '').strip() or None
        car.year = request.form.get('year') or None
        car.license_plate = request.form.get('license_plate', '').strip() or None
        db.session.commit()
        flash('Car updated.', 'success')
        return redirect(url_for('car_detail', car_id=car.id))
    return render_template('cars/form.html', car=car)


@app.route('/cars/<int:car_id>/delete', methods=['POST'])
@login_required
def car_delete(car_id):
    car = Car.query.get_or_404(car_id)
    if not current_user.is_admin and car.owner_id != current_user.id:
        abort(403)
    name = car.name
    db.session.delete(car)
    db.session.commit()
    flash(f'Car "{name}" deleted.', 'success')
    return redirect(url_for('cars_list'))


@app.route('/cars/<int:car_id>/share', methods=['GET', 'POST'])
@login_required
def car_share(car_id):
    car = Car.query.get_or_404(car_id)
    if not current_user.is_admin and car.owner_id != current_user.id:
        abort(403)
    if request.method == 'POST':
        action = request.form.get('action')
        if action == 'add':
            username = request.form.get('username', '').strip()
            user = User.query.filter_by(username=username).first()
            if not user:
                flash(f'User "{username}" not found.', 'danger')
            elif user.id == current_user.id:
                flash("You can't share with yourself.", 'warning')
            elif CarShare.query.filter_by(car_id=car_id, user_id=user.id).first():
                flash(f'Already shared with {user.username}.', 'warning')
            else:
                db.session.add(CarShare(car_id=car_id, user_id=user.id))
                db.session.commit()
                flash(f'Car shared with {user.username}.', 'success')
        elif action == 'remove':
            share = CarShare.query.filter_by(
                car_id=car_id,
                user_id=request.form.get('user_id', type=int),
            ).first_or_404()
            db.session.delete(share)
            db.session.commit()
            flash('Access removed.', 'success')
        return redirect(url_for('car_share', car_id=car_id))
    return render_template('cars/share.html', car=car,
                           shared_users=[s.user for s in car.shares])


# ─────────────────────────────────────────────────────────────────────────────
# Expenses
# ─────────────────────────────────────────────────────────────────────────────

def _parse_expense_form(form):
    """Parse common expense fields. Returns (data_dict, error_str)."""
    try:
        amount = float(form.get('amount', 0))
        if amount <= 0:
            return None, 'Amount must be greater than 0.'
    except (ValueError, TypeError):
        return None, 'Invalid amount.'

    currency = form.get('currency', 'CZK')
    if currency == 'CZK':
        amount_czk = amount
    else:
        rate = get_exchange_rate(currency, 'CZK')
        if rate is None:
            return None, f'Could not fetch exchange rate for {currency}. Please try again.'
        amount_czk = amount * rate

    try:
        exp_date = datetime.strptime(form.get('date', ''), '%Y-%m-%d').date()
    except ValueError:
        return None, 'Invalid date.'

    return {
        'amount': amount,
        'currency': currency,
        'amount_czk': amount_czk,
        'date': exp_date,
        'notes': form.get('notes', '').strip() or None,
    }, None


def _get_past_stations(user):
    """Return sorted unique fuel station names from all cars accessible to the user."""
    accessible_ids = [c.id for c in user.get_accessible_cars()]
    if not accessible_ids:
        return []
    rows = (db.session.query(FuelDetail.station_location)
            .join(Expense)
            .filter(Expense.car_id.in_(accessible_ids))
            .filter(FuelDetail.station_location.isnot(None))
            .distinct()
            .all())
    return sorted(r[0] for r in rows)


def _apply_type_detail(expense, form):
    """Create or update the type-specific detail row."""
    t = expense.expense_type

    if t == 'fuel':
        d = expense.fuel_detail or FuelDetail()
        d.expense_id = expense.id
        try:
            d.liters = float(form.get('liters') or 0) or None
            d.odometer = int(form.get('odometer') or 0) or None
            ppl = float(form.get('price_per_liter') or 0) or None
            if not ppl and d.liters and expense.amount:
                ppl = round(expense.amount / d.liters, 2)
            d.price_per_liter = ppl
        except (ValueError, TypeError):
            d.liters = d.price_per_liter = d.odometer = None
        d.station_location = form.get('station_location', '').strip() or None
        time_str = form.get('transaction_time', '').strip()
        try:
            d.transaction_time = datetime.strptime(time_str, '%H:%M').time() if time_str else None
        except ValueError:
            d.transaction_time = None
        if not expense.fuel_detail:
            db.session.add(d)

    elif t == 'repair':
        d = expense.repair_detail or RepairDetail()
        d.expense_id = expense.id
        d.description = form.get('description', '').strip() or None
        if not expense.repair_detail:
            db.session.add(d)

    elif t == 'toll':
        d = expense.toll_detail or TollDetail()
        d.expense_id = expense.id
        d.route = form.get('route', '').strip() or None
        d.toll_type = form.get('toll_type', '').strip() or None
        exp_str = form.get('toll_expiration_date', '').strip()
        try:
            new_expiry = datetime.strptime(exp_str, '%Y-%m-%d').date() if exp_str else None
        except ValueError:
            new_expiry = None
        if new_expiry != d.expiration_date:
            d.notification_sent = False
        d.expiration_date = new_expiry
        d.notify_before_expiry = bool(form.get('toll_notify_before_expiry'))
        try:
            d.notify_days_before = max(1, min(30, int(form.get('toll_notify_days_before') or 14)))
        except (ValueError, TypeError):
            d.notify_days_before = 14
        if not expense.toll_detail:
            db.session.add(d)

    elif t == 'insurance':
        d = expense.insurance_detail or InsuranceDetail()
        d.expense_id = expense.id
        d.insurance_type = form.get('insurance_type', '').strip() or None
        d.provider = form.get('provider', '').strip() or None
        exp_str = form.get('expiration_date', '').strip()
        try:
            new_expiry = datetime.strptime(exp_str, '%Y-%m-%d').date() if exp_str else None
        except ValueError:
            new_expiry = None
        if new_expiry != d.expiration_date:
            d.notification_sent = False
        d.expiration_date = new_expiry
        d.notify_before_expiry = bool(form.get('insurance_notify_before_expiry'))
        try:
            d.notify_days_before = max(1, min(30, int(form.get('insurance_notify_days_before') or 14)))
        except (ValueError, TypeError):
            d.notify_days_before = 14
        if not expense.insurance_detail:
            db.session.add(d)

    elif t == 'gadget':
        d = expense.gadget_detail or GadgetDetail()
        d.expense_id = expense.id
        d.gadget_type = form.get('gadget_type', '').strip() or None
        if not expense.gadget_detail:
            db.session.add(d)


@app.route('/cars/<int:car_id>/expenses/new', methods=['GET', 'POST'])
@login_required
def expense_new(car_id):
    car = Car.query.get_or_404(car_id)
    if not car.is_accessible_by(current_user):
        abort(403)

    expense_type = request.args.get('type', 'fuel')
    if expense_type not in EXPENSE_TYPES:
        expense_type = 'fuel'

    if request.method == 'POST':
        expense_type = request.form.get('expense_type', 'fuel')
        if expense_type not in EXPENSE_TYPES:
            expense_type = 'fuel'
        data, error = _parse_expense_form(request.form)
        if error:
            flash(error, 'danger')
        else:
            expense = Expense(car_id=car_id, user_id=current_user.id,
                              expense_type=expense_type, **data)
            db.session.add(expense)
            db.session.flush()
            _apply_type_detail(expense, request.form)
            db.session.commit()
            flash('Expense added.', 'success')
            return redirect(url_for('car_detail', car_id=car_id))

    past_stations = _get_past_stations(current_user)
    return render_template('expenses/form.html',
        car=car, expense=None, expense_type=expense_type,
        currencies=SUPPORTED_CURRENCIES, toll_types=TOLL_TYPES,
        insurance_types=INSURANCE_TYPES, today=date.today().isoformat(),
        past_stations=past_stations,
    )


@app.route('/expenses/<int:expense_id>/edit', methods=['GET', 'POST'])
@login_required
def expense_edit(expense_id):
    expense = Expense.query.get_or_404(expense_id)
    if not expense.can_edit(current_user):
        abort(403)
    car = expense.car

    if request.method == 'POST':
        data, error = _parse_expense_form(request.form)
        if error:
            flash(error, 'danger')
        else:
            expense.amount = data['amount']
            expense.currency = data['currency']
            expense.amount_czk = data['amount_czk']
            expense.date = data['date']
            expense.notes = data['notes']
            _apply_type_detail(expense, request.form)
            db.session.commit()
            flash('Expense updated.', 'success')
            return redirect(url_for('car_detail', car_id=car.id))

    past_stations = _get_past_stations(current_user)
    return render_template('expenses/form.html',
        car=car, expense=expense, expense_type=expense.expense_type,
        currencies=SUPPORTED_CURRENCIES, toll_types=TOLL_TYPES,
        insurance_types=INSURANCE_TYPES, today=date.today().isoformat(),
        past_stations=past_stations,
    )


@app.route('/expenses/<int:expense_id>/delete', methods=['POST'])
@login_required
def expense_delete(expense_id):
    expense = Expense.query.get_or_404(expense_id)
    if not expense.can_edit(current_user):
        abort(403)
    car_id = expense.car_id
    db.session.delete(expense)
    db.session.commit()
    flash('Expense deleted.', 'success')
    return redirect(url_for('car_detail', car_id=car_id))


# ─────────────────────────────────────────────────────────────────────────────
# Admin
# ─────────────────────────────────────────────────────────────────────────────

@app.route('/admin/overview')
@admin_required
def admin_overview():
    from collections import defaultdict
    users = User.query.filter_by(is_admin=False).order_by(User.username).all()
    cars  = Car.query.order_by(Car.name).all()

    sel_user_ids = request.args.getlist('user_id', type=int)
    sel_car_ids  = request.args.getlist('car_id',  type=int)
    sel_types    = [t for t in request.args.getlist('type') if t in EXPENSE_TYPES]

    q = Expense.query
    if sel_user_ids:
        q = q.filter(Expense.user_id.in_(sel_user_ids))
    if sel_car_ids:
        q = q.filter(Expense.car_id.in_(sel_car_ids))
    if sel_types:
        q = q.filter(Expense.expense_type.in_(sel_types))
    expenses = q.order_by(Expense.date.desc(), Expense.created_at.desc()).all()

    type_totals: dict = defaultdict(float)
    for e in expenses:
        type_totals[e.expense_type] += e.amount_czk

    # Build per-expense detail dict for the modal (serialised to JSON in template)
    expense_details = {}
    for e in expenses:
        d = {
            'date': e.date.strftime('%d.%m.%Y'),
            'car': e.car.name,
            'user': e.user.username,
            'type': e.expense_type,
            'type_label': TYPE_LABELS.get(e.expense_type, e.expense_type),
            'type_icon': TYPE_ICONS.get(e.expense_type, ''),
            'type_color': TYPE_COLORS.get(e.expense_type, '#6c757d'),
            'amount_czk': format_czk(e.amount_czk),
            'orig': f'{e.amount:,.2f} {e.currency}' if e.currency != 'CZK' else '',
            'notes': e.notes or '',
            'time': '',
            'fields': [],
        }
        if e.expense_type == 'fuel' and e.fuel_detail:
            fd = e.fuel_detail
            if fd.transaction_time:
                d['time'] = fd.transaction_time.strftime('%H:%M')
            fields = []
            if fd.liters:
                fields.append(['Liters', f'{fd.liters:.2f} L'])
            if fd.price_per_liter:
                fields.append(['Price/liter', f'{fd.price_per_liter:.2f} Kč'])
            if fd.odometer:
                fields.append(['Odometer', f'{fd.odometer:,} km'])
            if fd.station_location:
                fields.append(['Station', fd.station_location])
            d['fields'] = fields
        elif e.expense_type == 'repair' and e.repair_detail:
            if e.repair_detail.description:
                d['fields'] = [['Description', e.repair_detail.description]]
        elif e.expense_type == 'toll' and e.toll_detail:
            fields = []
            if e.toll_detail.route:
                fields.append(['Route', e.toll_detail.route])
            if e.toll_detail.toll_type:
                fields.append(['Type', e.toll_detail.toll_type])
            d['fields'] = fields
        elif e.expense_type == 'insurance' and e.insurance_detail:
            fields = []
            if e.insurance_detail.insurance_type:
                fields.append(['Type', e.insurance_detail.insurance_type])
            if e.insurance_detail.provider:
                fields.append(['Provider', e.insurance_detail.provider])
            if e.insurance_detail.expiration_date:
                fields.append(['Expires', e.insurance_detail.expiration_date.strftime('%d.%m.%Y')])
            d['fields'] = fields
        elif e.expense_type == 'gadget' and e.gadget_detail:
            if e.gadget_detail.gadget_type:
                d['fields'] = [['Item', e.gadget_detail.gadget_type]]
        expense_details[e.id] = d

    return render_template('admin/overview.html',
        users=users,
        cars=cars,
        expenses=expenses,
        sel_user_ids=sel_user_ids,
        sel_car_ids=sel_car_ids,
        sel_types=sel_types,
        total_czk=sum(e.amount_czk for e in expenses),
        type_totals=dict(type_totals),
        expense_details=expense_details,
    )


@app.route('/admin/overview/export')
@admin_required
def admin_overview_export():
    sel_user_ids = request.args.getlist('user_id', type=int)
    sel_car_ids  = request.args.getlist('car_id',  type=int)
    sel_types    = [t for t in request.args.getlist('type') if t in EXPENSE_TYPES]
    q = Expense.query
    if sel_user_ids:
        q = q.filter(Expense.user_id.in_(sel_user_ids))
    if sel_car_ids:
        q = q.filter(Expense.car_id.in_(sel_car_ids))
    if sel_types:
        q = q.filter(Expense.expense_type.in_(sel_types))
    expenses = q.order_by(Expense.date.desc(), Expense.created_at.desc()).all()
    return _export_xlsx(expenses, include_user=True, filename='expenses_export')


@app.route('/admin/users/<int:user_id>')
@admin_required
def admin_user_detail(user_id):
    from collections import defaultdict
    user = User.query.get_or_404(user_id)
    if user.is_admin:
        abort(404)
    owned_cars = Car.query.filter_by(owner_id=user_id).all()
    shared_car_ids = [s.car_id for s in CarShare.query.filter_by(user_id=user_id).all()]
    shared_cars = Car.query.filter(Car.id.in_(shared_car_ids)).all() if shared_car_ids else []
    expenses = (Expense.query.filter_by(user_id=user_id)
                .order_by(Expense.date.desc(), Expense.created_at.desc()).all())
    type_totals: dict = defaultdict(float)
    for e in expenses:
        type_totals[e.expense_type] += e.amount_czk
    return render_template('admin/user_detail.html',
        profile=user,
        owned_cars=owned_cars,
        shared_cars=shared_cars,
        expenses=expenses,
        total_czk=sum(e.amount_czk for e in expenses),
        type_totals=dict(type_totals),
    )


@app.route('/admin/requests/<int:user_id>/approve', methods=['POST'])
@admin_required
def admin_approve_username(user_id):
    user = User.query.get_or_404(user_id)
    if not user.pending_username:
        flash('No pending request.', 'warning')
        return redirect(url_for('admin_index'))
    if User.query.filter_by(username=user.pending_username).first():
        flash(f'Username "{user.pending_username}" is already taken. Request denied.', 'danger')
        user.pending_username = None
        user.pending_username_at = None
    else:
        flash(f'Username changed to "{user.pending_username}".', 'success')
        user.username = user.pending_username
        user.pending_username = None
        user.pending_username_at = None
    db.session.commit()
    return redirect(url_for('admin_index'))


@app.route('/admin/requests/<int:user_id>/deny', methods=['POST'])
@admin_required
def admin_deny_username(user_id):
    user = User.query.get_or_404(user_id)
    user.pending_username = None
    user.pending_username_at = None
    db.session.commit()
    flash('Username change request denied.', 'info')
    return redirect(url_for('admin_index'))


@app.route('/admin')
@admin_required
def admin_index():
    pending_users = User.query.filter(User.pending_username.isnot(None)).order_by(User.pending_username_at).all()
    users = User.query.order_by(User.username).all()
    return render_template('admin/index.html', users=users, pending_users=pending_users)


@app.route('/admin/users/new', methods=['GET', 'POST'])
@admin_required
def admin_user_new():
    if request.method == 'POST':
        username = request.form.get('username', '').strip()
        email = request.form.get('email', '').strip()
        password = request.form.get('password', '')
        is_admin = bool(request.form.get('is_admin'))

        errors = []
        if not username:
            errors.append('Username is required.')
        if not email:
            errors.append('Email is required.')
        if not password:
            errors.append('Password is required.')
        if username and User.query.filter_by(username=username).first():
            errors.append('Username already taken.')
        if email and User.query.filter_by(email=email).first():
            errors.append('Email already registered.')

        if errors:
            for e in errors:
                flash(e, 'danger')
        else:
            user = User(username=username, email=email, is_admin=is_admin)
            user.set_password(password)
            db.session.add(user)
            db.session.commit()
            flash(f'User "{username}" created.', 'success')
            return redirect(url_for('admin_index'))

    return render_template('admin/user_form.html', user=None)


@app.route('/admin/users/<int:user_id>/edit', methods=['GET', 'POST'])
@admin_required
def admin_user_edit(user_id):
    user = User.query.get_or_404(user_id)
    if request.method == 'POST':
        username = request.form.get('username', '').strip()
        email = request.form.get('email', '').strip()
        is_admin = bool(request.form.get('is_admin'))
        new_password = request.form.get('password', '').strip()

        errors = []
        if not username:
            errors.append('Username is required.')
        if not email:
            errors.append('Email is required.')
        existing_u = User.query.filter_by(username=username).first()
        if existing_u and existing_u.id != user.id:
            errors.append('Username already taken.')
        existing_e = User.query.filter_by(email=email).first()
        if existing_e and existing_e.id != user.id:
            errors.append('Email already registered.')

        if errors:
            for e in errors:
                flash(e, 'danger')
        else:
            user.username = username
            user.email = email
            user.is_admin = is_admin
            if new_password:
                user.set_password(new_password)
            db.session.commit()
            flash('User updated.', 'success')
            return redirect(url_for('admin_index'))

    return render_template('admin/user_form.html', user=user)


@app.route('/admin/users/<int:user_id>/delete', methods=['POST'])
@admin_required
def admin_user_delete(user_id):
    user = User.query.get_or_404(user_id)
    if user.id == current_user.id:
        flash("You can't delete your own account.", 'warning')
        return redirect(url_for('admin_index'))
    username = user.username
    db.session.delete(user)
    db.session.commit()
    flash(f'User "{username}" deleted.', 'success')
    return redirect(url_for('admin_index'))


# ─────────────────────────────────────────────────────────────────────────────
# API
# ─────────────────────────────────────────────────────────────────────────────

@app.route('/api/exchange-rate')
@login_required
def api_exchange_rate():
    from_currency = request.args.get('from', 'EUR')
    if from_currency not in SUPPORTED_CURRENCIES:
        return jsonify({'error': 'Unsupported currency'}), 400
    if from_currency == 'CZK':
        return jsonify({'rate': 1.0, 'from': 'CZK', 'to': 'CZK'})
    rate = get_exchange_rate(from_currency, 'CZK')
    if rate is None:
        return jsonify({'error': 'Could not fetch rate'}), 503
    return jsonify({'rate': rate, 'from': from_currency, 'to': 'CZK'})


# ─────────────────────────────────────────────────────────────────────────────
# CLI
# ─────────────────────────────────────────────────────────────────────────────

@app.cli.command('init-db')
def init_db_command():
    """Create all database tables."""
    db.create_all()
    click.echo('Database initialized.')


@app.cli.command('create-admin')
@click.argument('username')
@click.argument('email')
@click.argument('password')
def create_admin_command(username, email, password):
    """Create an admin user.  Usage: flask create-admin USERNAME EMAIL PASSWORD"""
    if User.query.filter_by(username=username).first():
        click.echo(f'Error: user "{username}" already exists.', err=True)
        return
    user = User(username=username, email=email, is_admin=True)
    user.set_password(password)
    db.session.add(user)
    db.session.commit()
    click.echo(f'Admin user "{username}" created.')


# DB is managed by Flask-Migrate. Run `flask db upgrade` to apply migrations.


# ─────────────────────────────────────────────────────────────────────────────
# Expiry notifications
# ─────────────────────────────────────────────────────────────────────────────

def _send_notification_email(user, expense_type, car_name, expense_date, expiry_date, days_left):
    try:
        msg = Message(
            subject=f'{expense_type} expiring in {days_left} days — {car_name}',
            recipients=[user.email],
            body=(
                f'Hello {user.username},\n\n'
                f'Your {expense_type.lower()} expense for {car_name} '
                f'(added on {expense_date.strftime("%d.%m.%Y")}) '
                f'expires on {expiry_date.strftime("%d.%m.%Y")} '
                f'({days_left} days from today).\n\n'
                f'Log in to your Car Expenses app to review or renew it.\n'
            ),
        )
        mail.send(msg)
    except Exception as e:
        app.logger.warning(f'Failed to send expiry notification to {user.email}: {e}')


def send_expiry_notifications():
    from datetime import timedelta
    with app.app_context():
        today = date.today()
        tolls = (TollDetail.query
                 .filter_by(notify_before_expiry=True, notification_sent=False)
                 .filter(TollDetail.expiration_date.isnot(None))
                 .all())
        for d in tolls:
            if d.expiration_date - today == timedelta(days=d.notify_days_before):
                _send_notification_email(
                    d.expense.user, 'Toll', d.expense.car.display_name,
                    d.expense.date, d.expiration_date, d.notify_days_before,
                )
                d.notification_sent = True

        insurances = (InsuranceDetail.query
                      .filter_by(notify_before_expiry=True, notification_sent=False)
                      .filter(InsuranceDetail.expiration_date.isnot(None))
                      .all())
        for d in insurances:
            if d.expiration_date - today == timedelta(days=d.notify_days_before):
                _send_notification_email(
                    d.expense.user, 'Insurance', d.expense.car.display_name,
                    d.expense.date, d.expiration_date, d.notify_days_before,
                )
                d.notification_sent = True

        db.session.commit()


# Start background scheduler (skip in Flask reloader child process to avoid duplicate)
if not (app.debug and not os.environ.get('WERKZEUG_RUN_MAIN')):
    from apscheduler.schedulers.background import BackgroundScheduler
    _scheduler = BackgroundScheduler()
    _scheduler.add_job(send_expiry_notifications, 'cron', hour=8, minute=0)
    _scheduler.start()


if __name__ == '__main__':
    app.run(debug=True)
